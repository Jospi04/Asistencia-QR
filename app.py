from flask import Flask, render_template, request, redirect, url_for, flash, jsonify, session, send_file
import os
from datetime import datetime
import pandas as pd
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import calendar

# Importar infraestructura
from src.infrastructure.mysql_connection import MySQLConnection
from src.infrastructure.repositories_mysql import (
    EmpresaRepositoryMySQL,
    EmpleadoRepositoryMySQL,
    AsistenciaRepositoryMySQL,
    HorarioEstandarRepositoryMySQL,
    EscaneoTrackingRepositoryMySQL
)

# Importar use cases
from src.use_cases.register_employee import RegisterEmployeeUseCase
from src.use_cases.mark_attendance import MarkAttendanceUseCase
from src.use_cases.list_companies import ListCompaniesUseCase
from src.use_cases.get_report import GetReportUseCase

# Importar QR generator
from src.infrastructure.qr_generator import QRGenerator

app = Flask(__name__)
app.secret_key = os.getenv('SECRET_KEY', 'tu_clave_secreta_aqui')

# Configuración de base de datos
db_connection = MySQLConnection()

# Inicializar repositorios
empresa_repo = EmpresaRepositoryMySQL(db_connection)
empleado_repo = EmpleadoRepositoryMySQL(db_connection)
asistencia_repo = AsistenciaRepositoryMySQL(db_connection)
horario_repo = HorarioEstandarRepositoryMySQL(db_connection)
escaneo_repo = EscaneoTrackingRepositoryMySQL(db_connection)

# Inicializar use cases
register_employee_use_case = RegisterEmployeeUseCase(empleado_repo)
mark_attendance_use_case = MarkAttendanceUseCase(empleado_repo, asistencia_repo, horario_repo, escaneo_repo)
list_companies_use_case = ListCompaniesUseCase(empresa_repo)
get_report_use_case = GetReportUseCase(empleado_repo, asistencia_repo, empresa_repo)

# Inicializar QR generator
qr_generator = QRGenerator()

def obtener_nombre_mes(numero_mes):
    """Obtiene el nombre del mes por su número"""
    meses = [
        '', 'Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio',
        'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre'
    ]
    return meses[numero_mes] if 1 <= numero_mes <= 12 else ''

# Rutas principales
@app.route('/')
def index():
    return render_template('index.html')

# Rutas de administración
@app.route('/admin')
def admin_dashboard():
    if not session.get('admin_logged_in'):
        return redirect(url_for('admin_login'))
    
    empresas = list_companies_use_case.execute()
    return render_template('admin_dashboard.html', empresas=empresas)

@app.route('/admin/login', methods=['GET', 'POST'])
def admin_login():
    if request.method == 'POST':
        # Aquí iría la lógica de autenticación
        username = request.form['username']
        password = request.form['password']
        
        # Por ahora, login simulado
        if username == 'admin' and password == 'admin':
            session['admin_logged_in'] = True
            return redirect(url_for('admin_dashboard'))
        else:
            flash('Credenciales inválidas', 'error')
    
    return render_template('login.html')

@app.route('/admin/logout')
def admin_logout():
    session.clear()
    return redirect(url_for('index'))

@app.route('/admin/add_employee', methods=['GET', 'POST'])
def admin_add_employee():
    if not session.get('admin_logged_in'):
        return redirect(url_for('admin_login'))
    
    if request.method == 'POST':
        try:
            nombre = request.form['nombre']
            empresa_id = int(request.form['empresa_id'])
            dni = request.form['dni']
            telefono = request.form.get('telefono', '')
            correo = request.form.get('correo', '')
            
            # Registrar empleado
            empleado = register_employee_use_case.execute(
                nombre=nombre,
                empresa_id=empresa_id,
                dni=dni,
                telefono=telefono,
                correo=correo
            )
            
            # Generar código QR para el empleado
            empresa = empresa_repo.get_by_id(empresa_id)
            if empresa:
                qr_path = qr_generator.generate_employee_qr(empleado.id, empresa.codigo_empresa)
                if qr_path:
                    flash('Empleado registrado con éxito. Código QR generado.', 'success')
                else:
                    flash('Empleado registrado, pero hubo un error generando el QR.', 'warning')
            else:
                flash('Empleado registrado con éxito.', 'success')
            
            return redirect(url_for('admin_add_employee'))
            
        except Exception as e:
            flash(f'Error registrando empleado: {str(e)}', 'error')
    
    # Obtener empresas para el formulario
    empresas = list_companies_use_case.execute()
    return render_template('admin_add_employee.html', empresas=empresas)

@app.route('/admin/employees')
def admin_list_employees():
    if not session.get('admin_logged_in'):
        return redirect(url_for('admin_login'))
    
    empresa_id = request.args.get('empresa_id', type=int)
    
    if empresa_id:
        empleados = empleado_repo.get_by_empresa_id(empresa_id)
        empresa = empresa_repo.get_by_id(empresa_id)
    else:
        empleados = empleado_repo.get_all()
        empresa = None
    
    empresas = list_companies_use_case.execute()
    return render_template('admin_list_employees.html', 
                         empleados=empleados, 
                         empresas=empresas, 
                         empresa_seleccionada=empresa)

# Ruta para editar empleado (formulario)
@app.route('/admin/edit_employee/<int:empleado_id>', methods=['GET', 'POST'])
def edit_employee(empleado_id):
    if not session.get('admin_logged_in'):
        return redirect(url_for('admin_login'))
    
    empleado = empleado_repo.get_by_id(empleado_id)
    if not empleado:
        flash('Empleado no encontrado', 'error')
        return redirect(url_for('admin_list_employees'))
    
    if request.method == 'POST':
        try:
            # Actualizar datos del empleado
            empleado.nombre = request.form['nombre']
            empleado.empresa_id = int(request.form['empresa_id'])
            empleado.dni = request.form['dni']
            empleado.telefono = request.form.get('telefono', '')
            empleado.correo = request.form.get('correo', '')
            
            # Guardar cambios
            empleado_repo.update(empleado)
            flash('Empleado actualizado con éxito', 'success')
            return redirect(url_for('admin_list_employees'))
            
        except Exception as e:
            flash(f'Error actualizando empleado: {str(e)}', 'error')
    
    # Obtener empresas para el formulario
    empresas = list_companies_use_case.execute()
    return render_template('admin_edit_employee.html', empleado=empleado, empresas=empresas)

# Ruta para activar/desactivar empleado (AJAX)
@app.route('/admin/toggle_employee/<int:empleado_id>', methods=['POST'])
def toggle_employee(empleado_id):
    if not session.get('admin_logged_in'):
        return jsonify({'success': False, 'message': 'No autorizado'}), 401
    
    try:
        empleado = empleado_repo.get_by_id(empleado_id)
        if not empleado:
            return jsonify({'success': False, 'message': 'Empleado no encontrado'}), 404
        
        # Alternar estado activo
        empleado.activo = not empleado.activo
        empleado_repo.update(empleado)
        
        estado = "activado" if empleado.activo else "desactivado"
        return jsonify({
            'success': True, 
            'message': f'Empleado {estado} con éxito',
            'activo': empleado.activo
        })
        
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error: {str(e)}'}), 500

# Ruta para eliminar empleado (AJAX) - ELIMINACIÓN COMPLETA
@app.route('/admin/delete_employee/<int:empleado_id>', methods=['POST'])
def delete_employee(empleado_id):
    if not session.get('admin_logged_in'):
        return jsonify({'success': False, 'message': 'No autorizado'}), 401
    
    try:
        empleado = empleado_repo.get_by_id(empleado_id)
        if not empleado:
            return jsonify({'success': False, 'message': 'Empleado no encontrado'}), 404
        
        nombre_empleado = empleado.nombre
        
        # Eliminar empleado COMPLETAMENTE de la base de datos
        empleado_repo.delete(empleado_id)
        
        return jsonify({
            'success': True, 
            'message': f'Empleado {nombre_empleado} eliminado permanentemente de la base de datos'
        })
        
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error: {str(e)}'}), 500

# Ruta para descargar QR de empleado
@app.route('/admin/download_qr/<int:empleado_id>')
def download_qr(empleado_id):
    try:
        # Obtener empleado
        empleado = empleado_repo.get_by_id(empleado_id)
        if not empleado:
            flash('Empleado no encontrado', 'error')
            return redirect(url_for('admin_list_employees'))
        
        # Obtener empresa
        empresa = empresa_repo.get_by_id(empleado.empresa_id)
        if not empresa:
            flash('Empresa no encontrada', 'error')
            return redirect(url_for('admin_list_employees'))
        
        # Generar código QR
        qr_path = qr_generator.generate_employee_qr(empleado.id, empresa.codigo_empresa)
        
        if qr_path and os.path.exists(qr_path):
            # Devolver el archivo como descarga
            return send_file(
                qr_path,
                mimetype='image/png',
                as_attachment=True,
                download_name=f'qr_empleado_{empleado_id}_{empresa.codigo_empresa}.png'
            )
        else:
            flash('Error generando código QR para descarga', 'error')
            return redirect(url_for('admin_list_employees'))
            
    except Exception as e:
        flash(f'Error descargando QR: {str(e)}', 'error')
        return redirect(url_for('admin_list_employees'))

# Rutas de escaneo
@app.route('/scan')
def scan_qr():
    return render_template('scan.html')

@app.route('/api/scan', methods=['POST'])
def api_scan_qr():
    try:
        data = request.get_json()
        codigo_qr = data.get('codigo_qr', '')
        ip_address = request.remote_addr
        
        # Procesar asistencia
        resultado = mark_attendance_use_case.execute(codigo_qr, ip_address)
        
        return jsonify(resultado)
        
    except Exception as e:
        return jsonify({
            "status": "error",
            "message": f"Error procesando escaneo: {str(e)}",
            "data": None
        })

# Rutas de reportes
@app.route('/reports')
def reports():
    empresas = list_companies_use_case.execute()
    return render_template('report.html', empresas=empresas)

@app.route('/api/reports/monthly')
def api_monthly_report():
    try:
        empresa_id = request.args.get('empresa_id', type=int)
        mes = request.args.get('mes', type=int, default=datetime.now().month)
        anio = request.args.get('anio', type=int, default=datetime.now().year)
        
        if not empresa_id:
            return jsonify({"error": "Empresa ID requerido"}), 400
        
        reporte = get_report_use_case.execute_monthly_report(empresa_id, mes, anio)
        return jsonify(reporte)
        
    except Exception as e:
        return jsonify({"error": f"Error generando reporte: {str(e)}"}), 500

@app.route('/api/reports/employee/<int:empleado_id>')
def api_employee_report(empleado_id):
    try:
        mes = request.args.get('mes', type=int, default=datetime.now().month)
        anio = request.args.get('anio', type=int, default=datetime.now().year)
        
        reporte = get_report_use_case.execute_employee_detail_report(empleado_id, mes, anio)
        return jsonify(reporte)
        
    except Exception as e:
        return jsonify({"error": f"Error generando reporte: {str(e)}"}), 500

# Ruta para exportar reporte a Excel
@app.route('/api/reports/export/excel')
def export_report_excel():
    try:
        empresa_id = request.args.get('empresa_id', type=int)
        mes = request.args.get('mes', type=int, default=datetime.now().month)
        anio = request.args.get('anio', type=int, default=datetime.now().year)
        
        if not empresa_id:
            flash('Empresa ID requerido', 'error')
            return redirect(url_for('reports'))
        
        # Obtener datos del reporte
        reporte = get_report_use_case.execute_monthly_report(empresa_id, mes, anio)
        
        if 'error' in reporte:
            flash(reporte['error'], 'error')
            return redirect(url_for('reports'))
        
        # Crear archivo Excel en memoria
        output = BytesIO()
        
        # Crear workbook
        wb = Workbook()
        
        # Hoja 1: Reporte de Empleados
        ws1 = wb.active
        ws1.title = "Reporte Empleados"
        
        # Encabezados
        headers = ['Nombre', 'DNI', 'Asistencias', 'Faltas', 'Horas Normales', 'Horas Extras', 'Porcentaje Asistencia (%)']
        ws1.append(headers)
        
        # Estilo para encabezados
        for col in range(1, len(headers) + 1):
            cell = ws1.cell(row=1, column=col)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Datos de empleados
        for empleado in reporte['empleados']:
            ws1.append([
                empleado['nombre'] or '',
                empleado['dni'] or '',
                empleado['asistencias'] or 0,
                empleado['faltas'] or 0,
                empleado['horas_normales'] or 0,
                empleado['horas_extras'] or 0,
                empleado['porcentaje_asistencia'] or 0
            ])
        
        # Ajustar ancho de columnas
        column_widths = [20, 15, 12, 8, 15, 15, 20]
        for i, width in enumerate(column_widths, 1):
            ws1.column_dimensions[chr(64 + i)].width = width
        
        # Hoja 2: Resumen
        ws2 = wb.create_sheet("Resumen")
        
        # Datos de resumen
        resumen_data = [
            ['Concepto', 'Valor'],
            ['Empresa', reporte['empresa']['nombre'] if reporte['empresa'] else ''],
            ['Período', f"{obtener_nombre_mes(mes)} {anio}"],
            ['Fecha Inicio', reporte['periodo']['primer_dia'] if reporte['periodo'] else ''],
            ['Fecha Fin', reporte['periodo']['ultimo_dia'] if reporte['periodo'] else ''],
            ['Días Laborables', reporte['totales']['dias_laborables'] if reporte['totales'] else 0],
            ['Total Empleados', reporte['totales']['total_empleados'] if reporte['totales'] else 0],
            ['Total Horas Normales', reporte['totales']['total_horas_normales'] if reporte['totales'] else 0],
            ['Total Horas Extras', reporte['totales']['total_horas_extras'] if reporte['totales'] else 0],
            ['Total Faltas', reporte['totales']['total_faltas'] if reporte['totales'] else 0],
            ['Generado', datetime.now().strftime('%d/%m/%Y %H:%M:%S')]
        ]
        
        for row in resumen_data:
            ws2.append(row)
        
        # Estilo para encabezados del resumen
        ws2.cell(row=1, column=1).font = Font(bold=True)
        ws2.cell(row=1, column=2).font = Font(bold=True)
        ws2.cell(row=1, column=1).fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        ws2.cell(row=1, column=2).fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Ajustar ancho de columnas del resumen
        ws2.column_dimensions['A'].width = 25
        ws2.column_dimensions['B'].width = 30
        
        # Guardar el workbook en el buffer
        wb.save(output)
        output.seek(0)
        
        # Preparar nombre de archivo
        nombre_empresa = reporte['empresa']['nombre'].replace(' ', '_') if reporte['empresa'] else 'empresa'
        nombre_archivo = f"reporte_asistencia_{nombre_empresa}_{mes}_{anio}.xlsx"
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=nombre_archivo
        )
        
    except Exception as e:
        flash(f'Error generando reporte Excel: {str(e)}', 'error')
        return redirect(url_for('reports'))

# Ruta para generar QR de empleado existente
@app.route('/admin/generate_qr/<int:empleado_id>')
def generate_employee_qr(empleado_id):
    if not session.get('admin_logged_in'):
        return redirect(url_for('admin_login'))
    
    try:
        empleado = empleado_repo.get_by_id(empleado_id)
        if not empleado:
            flash('Empleado no encontrado', 'error')
            return redirect(url_for('admin_list_employees'))
        
        empresa = empresa_repo.get_by_id(empleado.empresa_id)
        if not empresa:
            flash('Empresa no encontrada', 'error')
            return redirect(url_for('admin_list_employees'))
        
        # Generar QR
        qr_path = qr_generator.generate_employee_qr(empleado.id, empresa.codigo_empresa)
        if qr_path:
            flash('Código QR generado con éxito.', 'success')
        else:
            flash('Error generando código QR', 'error')
            
    except Exception as e:
        flash(f'Error: {str(e)}', 'error')
    
    return redirect(url_for('admin_list_employees'))

# Error handlers
@app.errorhandler(404)
def not_found(error):
    return render_template('error.html', error_message="Página no encontrada"), 404

@app.errorhandler(500)
def internal_error(error):
    return render_template('error.html', error_message="Error interno del servidor"), 500

if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5000)